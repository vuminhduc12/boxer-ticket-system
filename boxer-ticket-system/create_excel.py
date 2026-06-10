"""Generate ticket management Excel workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "ticket-management.xlsx"

RED = "B22222"
BLACK = "222222"
GOLD = "C9A84C"
WHITE = "FFFFFF"
GRAY_LIGHT = "F2F2F2"
GRAY_BORDER = "CCCCCC"

thin_border = Border(
    left=Side(style="thin", color=GRAY_BORDER),
    right=Side(style="thin", color=GRAY_BORDER),
    top=Side(style="thin", color=GRAY_BORDER),
    bottom=Side(style="thin", color=GRAY_BORDER),
)

title_fill = PatternFill("solid", fgColor=RED)
header_fill = PatternFill("solid", fgColor=BLACK)
white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor=GRAY_LIGHT)

title_font = Font(name="Arial", bold=True, size=14, color=WHITE)
header_font = Font(name="Arial", bold=True, size=11, color=GOLD)
body_font = Font(name="Arial", size=11)
section_font = Font(name="Arial", bold=True, size=11, color=RED)
bold_font = Font(name="Arial", bold=True, size=11)

HEADERS = [
    "No", "申込日", "お名前", "フリガナ", "メール", "電話番号",
    "種別", "枚数", "合計金額", "支払方法", "入金確認", "備考",
]

SAMPLE_DATA = [
    ["2025/06/01", "佐藤 健", "サトウ ケン", "sato@example.com", "090-1111-2222",
     "一般席", 2, "銀行振込", "✅入金済み", "2隣希望"],
    ["2025/06/02", "田中 美咲", "タナカ ミサキ", "tanaka@example.com", "080-3333-4444",
     "リングサイド", 1, "PayPay", "⏳未入金", ""],
    ["2025/06/03", "鈴木 グループ", "スズキ グループ", "suzuki@example.com", "070-5555-6666",
     "グループ", 5, "現金手渡し", "✅入金済み", "5名一括"],
]

COL_WIDTHS = [5, 12, 14, 14, 22, 14, 12, 8, 12, 12, 12, 20]


def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def create_application_sheet(wb):
    ws = wb.active
    ws.title = "申込管理"

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "🥊 応援チケット 申込管理シート"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    for i, row_data in enumerate(SAMPLE_DATA, 3):
        ws.cell(row=i, column=1, value=i - 2)
        for j, val in enumerate(row_data, 2):
            ws.cell(row=i, column=j, value=val)

        qty_cell = f"H{i}"
        type_cell = f"G{i}"
        ws.cell(row=i, column=9).value = (
            f'=IF({type_cell}="一般席",{qty_cell}*3000,'
            f'IF({type_cell}="リングサイド",{qty_cell}*5000,'
            f'IF({type_cell}="グループ",{qty_cell}*2500,0)))'
        )
        ws.cell(row=i, column=9).number_format = "#,##0"

        fill = white_fill if (i - 2) % 2 == 1 else gray_fill
        for col in range(1, 13):
            cell = ws.cell(row=i, column=col)
            cell.font = body_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    total_row = 6
    ws.cell(row=total_row, column=1, value="合計")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    ws.cell(row=total_row, column=1).font = bold_font
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=9, value="=SUM(I3:I5)")
    ws.cell(row=total_row, column=9).font = bold_font
    ws.cell(row=total_row, column=9).number_format = "#,##0"

    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    apply_border(ws, 1, total_row, 1, 12)
    ws.freeze_panes = "A3"


def create_fight_info_sheet(wb):
    ws = wb.create_sheet("試合情報")

    ws.merge_cells("A1:B1")
    ws["A1"].value = "🥊 試合・チケット基本情報"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    rows = [
        ("【試合情報】", ""),
        ("試合日", "2025年12月14日（日）"),
        ("会場", "○○アリーナ"),
        ("開場・開始時間", "開場 17:00 / 開始 18:00"),
        ("", ""),
        ("【チケット価格】", ""),
        ("一般席", "¥3,000（当日精算可）"),
        ("リングサイド", "¥5,000（要事前予約）"),
        ("グループ（5名以上）", "¥2,500/人（要事前申し込み）"),
        ("", ""),
        ("【銀行口座情報】", ""),
        ("銀行名", "○○銀行 ○○支店"),
        ("口座種別", "普通"),
        ("口座番号", "1234567"),
        ("口座名義", "ヤマダ タロウ"),
        ("", ""),
        ("【連絡先】", ""),
        ("電話", "090-XXXX-XXXX"),
        ("電話受付時間", "平日 20:00〜22:00 / 土日 10:00〜18:00"),
        ("メール", "your@email.com"),
        ("Instagram", "@yourhandle"),
        ("Facebook", "@yourpage"),
    ]

    for i, (label, value) in enumerate(rows, 2):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        if label.startswith("【"):
            label_cell.font = section_font
            value_cell.font = section_font
        else:
            label_cell.font = bold_font if label else body_font
            value_cell.font = body_font
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    apply_border(ws, 1, len(rows) + 1, 1, 2)


def create_dashboard_sheet(wb):
    ws = wb.create_sheet("集計ダッシュボード")

    ws.merge_cells("A1:B1")
    ws["A1"].value = "📊 チケット集計ダッシュボード"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    metrics = [
        ("総申込件数", "=COUNTA(申込管理!C3:C999)"),
        ("総チケット枚数", "=SUM(申込管理!H3:H999)"),
        ("合計売上", "=申込管理!I6"),
        ("入金済み件数", '=COUNTIF(申込管理!K3:K1000,"✅入金済み")'),
        ("未入金件数", '=COUNTIF(申込管理!K3:K1000,"⏳未入金")'),
        ("キャンセル件数", '=COUNTIF(申込管理!K3:K1000,"❌キャンセル")'),
        ("", ""),
        ("【種別ごとの枚数】", ""),
        ("一般席", '=SUMIF(申込管理!G3:G1000,"一般席",申込管理!H3:H1000)'),
        ("リングサイド", '=SUMIF(申込管理!G3:G1000,"リングサイド",申込管理!H3:H1000)'),
        ("グループ", '=SUMIF(申込管理!G3:G1000,"グループ",申込管理!H3:H1000)'),
        ("", ""),
        ("【支払方法ごとの件数】", ""),
        ("銀行振込", '=COUNTIF(申込管理!J3:J1000,"銀行振込")'),
        ("PayPay", '=COUNTIF(申込管理!J3:J1000,"PayPay")'),
        ("現金手渡し", '=COUNTIF(申込管理!J3:J1000,"現金手渡し")'),
    ]

    for i, (label, formula) in enumerate(metrics, 2):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=formula)
        if label.startswith("【"):
            label_cell.font = section_font
            value_cell.font = section_font
        else:
            label_cell.font = bold_font if label else body_font
            value_cell.font = body_font
        if label == "合計売上":
            value_cell.number_format = "¥#,##0"
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    apply_border(ws, 1, len(metrics) + 1, 1, 2)


def main():
    wb = Workbook()
    create_application_sheet(wb)
    create_fight_info_sheet(wb)
    create_dashboard_sheet(wb)
    wb.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
